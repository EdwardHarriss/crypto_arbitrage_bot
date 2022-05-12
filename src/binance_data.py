import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

import signal
import threading
import websocket
from datetime import datetime
import json
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from src.cex_data import CentralizedExchangeData as CEX_Data

def arbitrage(date, data):
    date = datetime.fromtimestamp(date/1000).strftime('%Y-%m-%d %H:%M:%S')
    BinanceData.GetArbitrageRoutes(data)
    output = BinanceData.GetArbitrageReturns(data, date)
    wb = load_workbook(filename="data/Triangular_Arbitrage_Binance.xlsx")
    ws = wb['Binance']
    for r in dataframe_to_rows(output, index=False, header=False):
        sr = ' -> '.join([str(elem) for elem in r[3]])
        r[3] = sr
        ws.append(r)
    
    ws = wb['Binance Occ.']
    for r in dataframe_to_rows(output, index=False, header=False):
        occ = [date, 'Binance', len(output)]
        ws.append(occ)
    wb.save("data/Triangular_Arbitrage_Binance.xlsx")

def on_message(ws, message):
    message = json.loads(message)

    def run(*args):
        arbitrage(message[0]['E'], {m['s']: float(m['c']) for m in message})

    threading.Thread(target=run).start()


def on_error(ws, error):
    print(error)


def on_close(ws, close_status_code, close_msg):
    print(close_status_code)
    print(close_msg)
    print("### closed ###")
    print("Trying to Reconnect")
    BinanceExchange(MIN_ARBITRAGE, FEE, BASE)

def handler(signum, frame):
    wb = load_workbook(filename="data/Triangular_Arbitrage_Binance.xlsx")
    ws = wb['Binance Timing']
    for r in dataframe_to_rows(BinanceData.timing_df, index=False, header=False):
        ws.append(r)
    wb.save("data/Triangular_Arbitrage_Binance.xlsx")
    print("### closed ###")
    exit(1)

signal.signal(signal.SIGINT, handler)

def BinanceExchange(minimum_arbitrage_allowance_perc, fee_per_transaction_percent, base_):
    global MIN_ARBITRAGE, FEE, BASE, BinanceData
    MIN_ARBITRAGE = minimum_arbitrage_allowance_perc
    FEE = fee_per_transaction_percent
    BASE = base_

    BinanceData = CEX_Data('Binance', base_, minimum_arbitrage_allowance_perc, fee_per_transaction_percent)

    socket_spot = 'wss://stream.binance.com/ws/!ticker@arr'
    ws = websocket.WebSocketApp(socket_spot,
                                on_message=on_message,
                                on_error=on_error,
                                on_close=on_close)

    ws.run_forever()