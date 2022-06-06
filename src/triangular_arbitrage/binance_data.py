import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

import signal
import sys
import threading
import websocket
from datetime import datetime
from datetime import timedelta
import json
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from src.triangular_arbitrage.cex_data import CentralizedExchangeData as CEX_Data
import ext.sql as sql
import pandas as pd
from dotenv import load_dotenv
import os
import pymysql
from sqlalchemy import create_engine

COUNT = 0

load_dotenv()
pw = os.getenv('pw')

CONN = create_engine("mysql+pymysql://" + "root" + ":" + pw + "@" + "localhost" + "/" + "tri_arb_data")


def arbitrage(date, data, quantity, volatility):
    global TIMING_TABLE
    date = datetime.fromtimestamp(date/1000).strftime('%Y-%m-%d %H:%M:%S')
    BinanceData.GetArbitrageRoutes(data)
    output_routes = BinanceData.GetArbitrageReturns(data, date, quantity, volatility)
    TIMING_TABLE = BinanceData.timing_df
    output_routes['route'] = [','.join(map(str, l)) for l in output_routes['route']]

    out_occ = pd.DataFrame({'date':[date],'occ':[len(output_routes)],'btc_vol':[abs(volatility['BTCUSDT'])]})

    output_routes.to_sql('arb_routes_real',con=CONN, if_exists='append',index=False)
    out_occ.to_sql('arb_occ_real',con=CONN, if_exists='append',index=False)

    #Excel writing
    """
    wb = load_workbook(filename="data/Triangular_Arbitrage_Binance.xlsx")
    ws = wb['Binance']
    for r in dataframe_to_rows(output, index=False, header=False):
        sr = ' -> '.join([str(elem) for elem in r[3]])
        r[3] = sr
        ws.append(r)

    ws = wb['Binance Occ.']
    for r in dataframe_to_rows(output, index=False, header=False):
        occ = [date, 'Binance', len(output), abs(volatility['BTCUSDT'])]
        ws.append(occ)
    wb.save("data/Triangular_Arbitrage_Binance.xlsx")
    """

    CheckTime(date)

def CheckTime(date):
    global COUNT
    COUNT = COUNT + 1
    if COUNT >= 600:
        COUNT = 0
        try:
            WriteTimingTable()
        except:
            print("Fail to write table")

def WriteTimingTable():

    print("Writing Table")

    opp_occ = []
    max_length = []
    average_time = []

    for col_name,data_df in TIMING_TABLE.iteritems():   #Finding times
        if col_name == 'times':
            values_list = data_df.values.tolist()
            for route in values_list:
                list_times = route.split(", ")
                opp_occ.append(len(list_times))
                last_time = None
                i = []
                j = 0
                for x in list_times:
                    if last_time is None:
                        j = j+1
                        last_time = datetime.strptime(x, '%Y-%m-%d %H:%M:%S')
                    elif (datetime.strptime(x, '%Y-%m-%d %H:%M:%S')-timedelta(0,1)) == last_time :
                        j = j+1
                        last_time = datetime.strptime(x, '%Y-%m-%d %H:%M:%S')
                    else: 
                        j = 1
                    i.append(j)
                max_length.append(max(i))
                average_time.append(sum(i)/len(i))


    tmp_df = TIMING_TABLE
    tmp_df['max'] = max_length
    tmp_df['ave'] = average_time
    tmp_df['occ'] = opp_occ

    tmp_df = tmp_df.drop('times', 1)

    tmp_df.to_sql('arb_timing_real_11',con=CONN,if_exists ='replace',index=False)

"""
    wb = load_workbook(filename="data/Triangular_Arbitrage_Binance.xlsx")
    ws = wb['Binance Timing']
    for r in dataframe_to_rows(tmp_df, index=False, header=False):
        ws.append(r)
    wb.save("data/Triangular_Arbitrage_Binance.xlsx")
    print("### closed ###")
"""

def on_message(ws, message):
    message = json.loads(message)

    def run(*args):
        print(len(message.index))
        #arbitrage(message[0]['E'], {m['s']: float(m['c']) for m in message}, {m['s']: float(m['v']) for m in message}, {m['s']: float(m['P']) for m in message})

    threading.Thread(target=run).start()


def on_error(ws, error):
    print(error)


def on_close(ws, close_status_code, close_msg):
    print(close_status_code)
    print(close_msg)
    print("### closed ###")
    print("Trying to Reconnect")
    BinanceExchange(MIN_ARBITRAGE, FEE, BASE, TIMING_TABLE)

def handler(sig, frame):
    try:
        WriteTimingTable()
    except:
        print("Error in writing table")
    print("Closing")
    sys.exit(0)

signal.signal(signal.SIGINT, handler)

def BinanceExchange(minimum_arbitrage_allowance_perc, fee_per_transaction_percent, base_, TIMING_TABLE):
    global MIN_ARBITRAGE, FEE, BASE, BinanceData
    MIN_ARBITRAGE = minimum_arbitrage_allowance_perc
    FEE = fee_per_transaction_percent
    BASE = base_

    BinanceData = CEX_Data('Binance', base_, minimum_arbitrage_allowance_perc, fee_per_transaction_percent, TIMING_TABLE)

    socket_spot = 'wss://stream.binance.com/ws/!ticker@arr'
    ws = websocket.WebSocketApp(socket_spot,
                                on_message=on_message,
                                on_error=on_error,
                                on_close=on_close)

    ws.run_forever()